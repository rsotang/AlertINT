
import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import pandas as pd
import warnings

#esta linea esta para que no se vea el mensaje de error de la libreria pandas al limpiar los datos duplicados del biplano
pd.options.mode.chained_assignment = None  # default='warn'
warnings.simplefilter(action='ignore', category=FutureWarning)
#ESTA ES UNA VERSION MODIFICA DE import_XML_biplano.py CON EL FIN DE VERIFICAR SI FUNCIONA IGUAL DE BIEN PARA LOS Equipos e SIemens Y MÁS ADELANTE AÑADIR REDUNDANCIA Y CONTROL DE ERRORES AL SCRIPT COMPLEO

###################################################

def parse_xml_to_df(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Crear una lista para almacenar los datos
    data_list = []

    # Verificar si se encontró <Query_Criteria>
    query_criteria = root.find('Query_Criteria')
    if query_criteria is not None:
        base_data = query_criteria.attrib
        #print("Query_Criteria encontrado:", base_data)

        # Recorrer cada <DoseInfo> y sus subelementos
        for dose_info in query_criteria.findall('DoseInfo'):
            dose_info_data = dose_info.attrib
            #print("DoseInfo encontrado:", dose_info_data)

            # Extraer <Observer_Context> asociado
            observer_context = dose_info.find('Observer_Context')
            if observer_context is not None:
                dose_info_data.update(observer_context.attrib)
                #print("Observer_Context encontrado:", observer_context.attrib)

            # Extraer todos los <CT_Accumulated_Dose_Data>
            for accumulated_dose in dose_info.findall('CT_Accumulated_Dose_Data'):
                accumulated_data = accumulated_dose.attrib
                #print("CT_Accumulated_Dose_Data encontrado:", accumulated_data)
                combined_data = {**base_data, **dose_info_data, **accumulated_data}
                data_list.append(combined_data)

            # Extraer todos los <CT_Acquisition>
            #for acquisition in dose_info.findall('CT_Acquisition'):
            #    acquisition_data = acquisition.attrib
                #print("CT_Acquisition encontrado:", acquisition_data)
            #    combined_data = {**base_data, **dose_info_data, **acquisition_data}
            #    data_list.append(combined_data)
    #Esta es la parte que contiene la información mecánica de la maquina y que ene este caso nos da un poco igual
    else:
        print("No se encontró <Query_Criteria> en", xml_file)

    # Convertir la lista de diccionarios a un DataFrame
    if data_list:
        df = pd.DataFrame(data_list)
    else:
        df = pd.DataFrame()  # DataFrame vacío en caso de no encontrar datos

    return df

def ImportXML(input_dir):
    # Leer y combinar todos los archivos XML en un solo DataFrame
    all_data = pd.DataFrame()

    for file in os.listdir(input_dir):
        if file.endswith('.xml'):
            file_path = os.path.join(input_dir, file)
            print("Procesando archivo:", file_path)
            df = parse_xml_to_df(file_path)
            all_data = pd.concat([all_data, df], ignore_index=True)

    return all_data
    
    
#Esta función esta para eliminar las unidades de los xmls y poder quedarnos solo con los numeros

def clean_and_convert_to_float(column):
    # Extrae solo la parte numérica antes de cualquier carácter no numérico y convierte a float
    return pd.to_numeric(column.str.extract(r'^(\d*\.?\d*)')[0], errors='coerce')
    
########################################################################################

def main():
    
    input_dir = 'Pruebas de Victor/Paciente prueba Alertint/XML Hibrido'
    all_data = ImportXML(input_dir)
    #print(all_data)
    
    #limpiar los datos del xml y convertirlos a float
    
    if 'Total_Acquisition_Time' in all_data.columns:
        all_data['Total_Acquisition_Time'] = clean_and_convert_to_float(all_data['Total_Acquisition_Time'])
    else:
        all_data['Total_Acquisition_Time'] = 0

    if 'Total_Fluoro_Time' in all_data.columns:
        all_data['Total_Fluoro_Time'] = clean_and_convert_to_float(all_data['Total_Fluoro_Time'])
    else:
        all_data['Total_Fluoro_Time'] = 0
    #unidades el PDA a Gy/cm2 
    all_data['Dose_Area_Product_Total'] = clean_and_convert_to_float(all_data['Dose_Area_Product_Total']) * 10000

    all_data['Dose_RP_Total'] = clean_and_convert_to_float(all_data['Dose_RP_Total'])

    # Crear el campo "Tiempo de intervención y ponerlo en minutos"
    #all_data['Tiempo de intervención'] = (all_data['Total_Acquisition_Time'] + all_data['Total_Fluoro_Time']) / 60
    all_data['Tiempo de intervención'] = all_data['Total_Fluoro_Time'] / 60

    # Añadir campos adicionales y predefinir valores
    all_data['equipo'] = 'HIBRIDO'
    all_data['servicio'] = 'CIRUGIA VASCULAR'
    all_data['nombre paciente'] = ''
    all_data['Seguimiento'] = ''

    # Reordenar columnas según el orden especificado
    columns_order = [
        'equipo', 'servicio', 'PatientID', 'nombre paciente', 'SeriesDate', 
        'SeriesTime', 'StudyDescription', 'Dose_Area_Product_Total', 
        'Dose_RP_Total', 'Tiempo de intervención', 'Seguimiento'
    ]
    all_data = all_data.reindex(columns=columns_order)
    #print('punto1')
    #print(all_data)     
    

    #Condiciones de seguimiento
    ######################################
    #################################
    umbral_PDA = 500 #Gycm2
    umbral_DPR = 5 #Gy
    umbral_Tiempo = 60 #min

    ################################
    #######################################
    all_data['Seguimiento'] = all_data.apply(lambda row: 'SI' if row['Dose_Area_Product_Total'] > umbral_PDA or row['Dose_RP_Total'] > umbral_DPR or row['Tiempo de intervención'] > umbral_Tiempo else 'NO', axis=1) 
    print(all_data)
    
    #Generamos array secundario para seleccionar los pacientes con seguimieno
    df = all_data[all_data['Seguimiento'] == 'SI']
# Guardar todos los datos procesados
    #all_data.to_csv('Pacienteshibrido.csv', index=False)
    #all_data.to_excel('Pacienteshibrido.xlsx', index=False)

    ###############################################################
    print('Limpiando formatos para el excel')
    print(df)
    if df.empty:
        print('No se encontraron pacientes que necesiten seguimiento')
        exit()

    df['SeriesDate'] = pd.to_datetime(df['SeriesDate'], format='%Y%m%d').dt.date
    df['SeriesTime'] = pd.to_datetime(df['SeriesTime'], format='%H%M%S.%f').dt.time

    print('moviendo al excel')
    archivo_excel = 'Pruebas de Victor/pacientesprueba.xlsm'

    # Cargar el libro de trabajo existente
    libro = load_workbook(archivo_excel, keep_vba=True)

    # Seleccionar la hoja de trabajo en la que quieres añadir los datos
    hoja = libro['Pendientes']

    # Encontrar la última fila con contenido en la hoja
    ultima_fila = 1
    for fila in hoja.iter_rows(min_row=1, max_col=1, values_only=True):
       if all(cell is None for cell in fila):
           break
       ultima_fila += 1

    # Añadir las filas del DataFrame a partir de la última fila con contenido
    for i, fila in df.iterrows():
       for j, valor in enumerate(fila):
           celda = hoja.cell(row=ultima_fila + i , column=j + 1, value=valor)
            # Centrar el contenido de la celda
           celda.alignment = Alignment(horizontal='center', vertical='center')
           # Aplicar el formato adecuado si es fecha u hora
           if j == 4:  # Suponiendo que la cuara columna es la fecha
               celda.number_format = 'DD/MM/YYYY'  # Formato de fecha
           elif j == 5:  # Suponiendo que la quinta columna es la hora
               celda.number_format = 'HH:MM'  # Formato de hora


    # Guardar los cambios en el archivo de Excel
    libro.save(archivo_excel)

    print("Datos añadidos exitosamente.")



if __name__ == "__main__":
    main()





