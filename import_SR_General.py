#ESTA ES UNA VERSION REPETIDA DEL SCRIPT SR_PHILLIPS.PY QUE ES EL ORIGINAL DONDE ESTÁN TODOS LOS COMENTARIOS

#De momento parece que para General y Phillips los campos de los que sacamos los datos coinciden, pero puede ser susceptible de error, asi que en caso de que no se importen bien los datos de una maquina solo un posible erro rpuede ser que hayan actualizado los campos del SR de esa maquina y que no lo coja bien el script.

import pydicom
import pandas as pd
import os
from datetime import datetime, timedelta
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import logging
import traceback
################################################################################################################################################

# Función para extraer la información de dosis de un archivo DICOM SR
def extraer_datos_dosis(archivo_dicom):
    ds = pydicom.dcmread(archivo_dicom)
    
    dosis_total = 0.0
    tiempo_total_intervencion = 0.0
    pda_total = 0.0
    tiempo_fluo = 0.0  # Datos separados que se suman manualmente
    tiempo_adq = 0.0
    
    for item in ds.ContentSequence:
        if hasattr(item, 'ContentSequence'):
            for subitem in item.ContentSequence:
                if subitem.ConceptNameCodeSequence[0].CodeValue == '113725':  # Dosis a punto de referencia
                    dosis_referencia = float(subitem.MeasuredValueSequence[0].NumericValue)
                    dosis_total += dosis_referencia
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113730':  # Tiempo de fluoroscopia
                    tiempo_fluo = float(subitem.MeasuredValueSequence[0].NumericValue)
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113855':  # Tiempo de adquisición
                    tiempo_adq = float(subitem.MeasuredValueSequence[0].NumericValue)
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113722':  # Producto dosis-área (PDA)
                    pda = float(subitem.MeasuredValueSequence[0].NumericValue)
                    pda_total += pda * 10000

    tiempo_total_intervencion = tiempo_fluo / 60  # Convertimos a minutos
    return dosis_total, tiempo_total_intervencion, pda_total
################################################################################################################################################


# Función para procesar los archivos DICOM y extraer los datos de cada paciente
def procesar_archivos(carpeta_srdicom):
    datos_pacientes = []
    # Obtener la fecha del día anterior
    fecha_ayer = datetime.now() - timedelta(days=1)
    fecha_ayer = fecha_ayer.date()
    # Iterar sobre los archivos en la carpeta
    for archivo in os.listdir(carpeta_srdicom):
        if archivo.endswith('.dcm'):
            archivo_dicom = os.path.join(carpeta_srdicom, archivo)

            # Obtener la fecha de modificación del archivo
            fecha_modificacion = datetime.fromtimestamp(os.path.getmtime(archivo_dicom)).date()

            # Solo procesar si el archivo fue modificado el día anterior
            if fecha_modificacion == fecha_ayer:
                ds = pydicom.dcmread(archivo_dicom)
                # Extraer información adicional del paciente y estudio
                id_paciente = ds.PatientID
                nombre_paciente = str(ds.PatientName)
                fecha_intervencion = ds.StudyDate
                hora_intervencion = ds.StudyTime
                descripcion_estudio = ds.StudyDescription if 'StudyDescription' in ds else 'N/A'
                
                # Extraer información de dosis, tiempo y PDA
                dosis_pref, tiempo_total_intervencion, pda_total = extraer_datos_dosis(archivo_dicom)
                
                # Añadir los datos al DataFrame
                datos_pacientes.append({
                    'equipo': 'GE Vascular',
                    'servicio': 'Cardiología',
                    'PatientID': id_paciente,
                    'nombre paciente': nombre_paciente,
                    'SeriesDate': fecha_intervencion,
                    'SeriesTime': hora_intervencion,
                    'StudyDescription': descripcion_estudio,
                    'Dose_Area_Product_Total': pda_total,
                    'Dose_RP_Total': dosis_pref,
                    'Tiempo de intervención': tiempo_total_intervencion, 
                    'Seguimiento': ''
                })
    if not datos_pacientes:
        print('No se han encontrado archivos generados ayer, ', fecha_ayer)
    # Crear un DataFrame con los datos recolectados
    return pd.DataFrame(datos_pacientes)
################################################################################################################################################

# Función para limpiar los datos y aplicar las reglas de seguimiento
def limpiar_y_aplicar_seguimiento(df, umbral_PDA, umbral_DPR, umbral_Tiempo):
 
        
    df['SeriesDate'] = pd.to_datetime(df['SeriesDate'], format='%Y%m%d').dt.date
    df['SeriesTime'] = pd.to_datetime(df['SeriesTime'], format='%H%M%S.%f').dt.time
    df['Dose_RP_Total']= df['Dose_RP_Total'].round(2)
    df['Dose_Area_Product_Total']=df['Dose_Area_Product_Total'].round(2)
    df['Tiempo de intervención']=df['Tiempo de intervención'].round(2)
    
    df['Seguimiento'] = df.apply(lambda row: 'SI' if row['Dose_Area_Product_Total'] > umbral_PDA or row['Dose_RP_Total'] > umbral_DPR or row['Tiempo de intervención'] > umbral_Tiempo else 'NO', axis=1)
    
    result_df = df[df['Seguimiento'] == 'SI'].reset_index(drop=True)
    
    if result_df.empty:
        print('No se encontraron pacientes que necesiten seguimiento')
        sys.exit()

    return result_df
################################################################################################################################################

# Función para exportar los datos al archivo de Excel
def exportar_a_excel(df, archivo_excel, hoja_nombre='Pendientes'):
    
    print('Limpiando y añadiendo datos a hoja Excel: '+ archivo_excel)
    
    libro = load_workbook(archivo_excel, keep_vba=True)
    hoja_pendientes = libro[hoja_nombre]
    
    # Encontrar la última fila con contenido
    ultima_fila_pendientes = 1
    for fila in hoja_pendientes.iter_rows(min_row=1, max_col=1, values_only=True):
        if fila[0] is None:
            break
        ultima_fila_pendientes += 1

    # Añadir las filas del DataFrame a partir de la última fila
    for i, fila in df.iterrows():
        for j, valor in enumerate(fila):
            celda = hoja_pendientes.cell(row=ultima_fila_pendientes + i, column=j + 1, value=valor)
            celda.alignment = Alignment(horizontal='center', vertical='center')
            if j == 4:  # Columna de fecha
                celda.number_format = 'DD/MM/YYYY'
            elif j == 5:  # Columna de hora
                celda.number_format = 'HH:MM'

    libro.save(archivo_excel)
    print("Datos añadidos exitosamente.")
################################################################################################################################################

##########################################
##########################################

################################################################################################################################################
# Función principal
def main():
     #fragmento inicial para control de errores
    # Obtener la fecha actual y formatearla para incluirla en el nombre del archivo
    current_date = datetime.now().strftime("%Y-%m-%d")
    log_filename = f'Z:\Dosis pacientes\Alertas_Intervencionismo\_PRUEBAS EN CURSO-AlertInt python\_Logs errores\HemoGeneral\error_log_{current_date}.txt'
    
    # Crear un logger
    logger = logging.getLogger()
    logger.setLevel(logging.ERROR)
    
    # Crear un manejador de archivo que escriba en el archivo log
    file_handler = logging.FileHandler(log_filename)
    file_handler.setLevel(logging.ERROR)
    
    # Definir el formato del log
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    
    # Añadir el manejador al logger
    logger.addHandler(file_handler)

############################################################################
    try:  #ESTE ES EL CODIGO DE VERDAD  .    
       
        carpeta_srdicom = 'Z:\Dosis pacientes\HCUV Angiografo GE\Ficheros SR'
        archivo_excel = 'Z:\Dosis pacientes\Alertas_Intervencionismo\_PRUEBAS EN CURSO-AlertInt python\ppacientes.xlsm'
        print('Buscando pacientes de General en ' + carpeta_srdicom)
        umbral_PDA=500
        umbral_DPR=5
        umbral_Tiempo=60
        
        print('Procesando archivos DICOM...')
        df = procesar_archivos(carpeta_srdicom)
        
        
        df_limpio = limpiar_y_aplicar_seguimiento(df, umbral_PDA, umbral_DPR,umbral_Tiempo)
        print('Limpiando datos y aplicando reglas de seguimiento...')
        print(df_limpio)
        print('Exportando los datos al archivo de Excel...')
        exportar_a_excel(df_limpio, archivo_excel)
    
##############################################################################
        
    except Exception as e:
        # Capturar y registrar el error
        logger.error("Error occurred: %s", e)
        logger.error(traceback.format_exc())
    
    finally:
        # Asegurarse de que se cierra el manejador del archivo
        logger.removeHandler(file_handler)
        file_handler.close()


if __name__ == "__main__":
    main()