import os
from datetime import datetime, timedelta
import sys
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import warnings
import logging
import traceback

#esta linea esta para que no se vea el mensaje de error de la libreria pandas al limpiar los datos duplicados del biplano
pd.options.mode.chained_assignment = None  # default='warn'
warnings.simplefilter(action='ignore', category=FutureWarning)


#Esta función intenta leer un archivo XML y devolver un DataFrame con los datos extraídos
#la forma de las funciones esta adecuada a la estructura de los XML del Biplano en concreto
#de acuerdo a la forma de nodos XML padre e hijos.
#
#Para el Biplano se generan archivos XML con todos los pacientes que se han hecho en cada dia
#Esto es por un query automatizado diario.
# 
#El XML tiene una estructura un poco rara por eso mismo.
#Dentro de root hay un único hijo que es <Query> que es en verdad el root que nos interesa
#dentro de Query todos los hijos son de etiqueta <Doseinfo>, que se corresponde con cada paciente
#Dentro de <DoseInfo> (es decir de cada paciente) estan anidados en paralelo los nodos hijos de 
#<ObserverContext>, <CT_Accumulated_Dose_Data> (este nodo está dos veces porque genera uno por panel, por eso de que es un arco biplano y tal), y <CT_Acquisition> (este nodo está tantas veces como veces hayan pisado el pedal de fluoro o cine)
#De momento no parece que haya que modificar el script para adaptarlo al hibrido o a cualquier maquina de Siemenes

###################################################

def parse_xml_to_df(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Crear una lista para almacenar los datos
    data_list = []

    # Verificar si se encontró <Query_Criteria>
    query_criteria = root.find('Query_Criteria')
    if query_criteria is not None:
        base_data = query_criteria.attrib
        #print("Query_Criteria encontrado:", base_data)

        # Recorrer cada <DoseInfo> y sus subelementos
        for dose_info in query_criteria.findall('DoseInfo'):
            dose_info_data = dose_info.attrib
            #print("DoseInfo encontrado:", dose_info_data)

            # Extraer <Observer_Context> asociado
            observer_context = dose_info.find('Observer_Context')
            if observer_context is not None:
                dose_info_data.update(observer_context.attrib)
                #print("Observer_Context encontrado:", observer_context.attrib)

            # Extraer todos los <CT_Accumulated_Dose_Data>
            for accumulated_dose in dose_info.findall('CT_Accumulated_Dose_Data'):
                accumulated_data = accumulated_dose.attrib
                #print("CT_Accumulated_Dose_Data encontrado:", accumulated_data)
                combined_data = {**base_data, **dose_info_data, **accumulated_data}
                data_list.append(combined_data)

            # Extraer todos los <CT_Acquisition> 
#            for acquisition in dose_info.findall('CT_Acquisition'):
 #               acquisition_data = acquisition.attrib
 #               #print("CT_Acquisition encontrado:", acquisition_data)
 #               combined_data = {**base_data, **dose_info_data, **acquisition_data}
 #               data_list.append(combined_data)
#Esta es la parte que contiene la información mecánica de la maquina y que ene este caso nos da un poco igual
    
    else:
        print("No se encontró <Query_Criteria> en", xml_file)

    # Convertir la lista de diccionarios a un DataFrame
    return pd.DataFrame(data_list) if data_list else pd.DataFrame()

################################################################################################################################################
# Leer y combinar todos los archivos XML generados por careanalytics en un solo DataFrame

def import_xml(input_dir):

# Obtener la fecha de ayer
    yesterday = datetime.now().date() - timedelta(days=1)
    all_data = pd.DataFrame()

    for file in os.listdir(input_dir):
        if file.endswith('.xml'):
            file_path = os.path.join(input_dir, file)
             # Obtener la fecha de modificación del archivo
            file_mod_time = datetime.fromtimestamp(os.path.getmtime(file_path)).date()
             # Procesar solo los archivos modificados el día anterior
            if file_mod_time == yesterday:
                print("Procesando archivo:", file_path)
                df = parse_xml_to_df(file_path)
                all_data = pd.concat([all_data, df], ignore_index=True)
          
                 
    if all_data.empty:
        print('No se han encontrado archivos generados ayer, ', yesterday)
         
                
    return all_data
################################################################################################################################################
#Esta función esta para eliminar las unidades de los xmls y poder quedarnos solo con los numeros

def clean_and_convert_to_float(column):
        # Extrae solo la parte numérica antes de cualquier carácter no numérico y convierte a float
    return pd.to_numeric(column.str.extract(r'^(\d*\.?\d*)')[0], errors='coerce')
################################################################################################################################################


def remove_zero_rows(df):
    return df[(df['Dose_Area_Product_Total'] > 0) & (df['Dose_RP_Total'] > 0)].reset_index(drop=True)

################################################################################################################################################
def transform_data(df): #pilla las colummnas que nos interesan y las limpia
    df['Total_Acquisition_Time'] = clean_and_convert_to_float(df['Total_Acquisition_Time']) if 'Total_Acquisition_Time' in df.columns else 0
    df['Total_Fluoro_Time'] = clean_and_convert_to_float(df['Total_Fluoro_Time']) if 'Total_Fluoro_Time' in df.columns else 0
    df['Dose_Area_Product_Total'] = clean_and_convert_to_float(df['Dose_Area_Product_Total']) * 10000
    df['Dose_RP_Total'] = clean_and_convert_to_float(df['Dose_RP_Total'])
    df['Tiempo de intervención'] = df['Total_Fluoro_Time'] / 60 
    #cambio de unidades a minutos y Gy/cm2  y limpieza de caracteres
    
    #asignacion de los valores que nos interesan en ciertas columnas para poder hacer avisos en informes en el Excel
    df['equipo'] = 'BIPLANO'
    df['servicio'] = 'Neurointervencionismo'
    df['nombre paciente'] = ''
    df['Seguimiento'] = ''

    columns_order = [
        'equipo', 'servicio', 'PatientID', 'nombre paciente', 'SeriesDate', 
        'SeriesTime', 'StudyDescription', 'Dose_Area_Product_Total', 
        'Dose_RP_Total', 'Tiempo de intervención', 'Seguimiento'
    ]
    return df.reindex(columns=columns_order) #limpianodo los indices vacios de la matriz que quedan por reordenar sus elementos (movidas de Python). Si no se hace, quedan filas en blanco haciendo huecos entre los datos pegados
################################################################################################################################################
#FUNCION DE RESERVA POR SI DEJA DE FUNCIONAR ILOC EN UNA VERSION NUEVA DE PANDAS

def aggregate_patient_data(df):
    # Agrupar por 'PatientID' y sumar las columnas numéricas para juntar los pacientes que aparecen por duplicado (uno por panel)
    aggregated_df = df.groupby('PatientID', as_index=False).agg({
        'SeriesDate': 'first',  # Puede usar 'first' o 'min' si la fecha es la misma para todas las entradas
        'SeriesTime': 'first',  # Similar a la fecha, depende de su uso
        'StudyDescription': 'first',
        'Tiempo de intervención': 'first', #SOLO COGEMOS EL TUBO A
        'Dose_Area_Product_Total': 'sum',
        'Dose_RP_Total': 'sum'
    })
    return aggregated_df
################################################################################################################################################

def sum_patient_data(df): #sumamos los pacientes duplicados aprovechandonos de que siempre aparecen apareados en la lista de pacientes del xml
    result_df = pd.DataFrame(columns=df.columns)
     # Iterar por las filas
    for i in range(len(df)):
        if i == 0 or df.loc[i, 'PatientID'] != df.loc[i-1, 'PatientID']:
             # Si es la primera fila o el valor de 'PatientID es diferente al anterior, agregar la fila al resultado
            result_df = pd.concat([result_df, df.loc[[i]]], ignore_index=True)
        else:
              # Si el valor es igual al anterior, sumar los valores a la última fila del resultado
            result_df.loc[result_df.index[-1], 'Dose_Area_Product_Total'] += df.loc[i, 'Dose_Area_Product_Total']
            result_df.loc[result_df.index[-1], 'Dose_RP_Total'] += df.loc[i, 'Dose_RP_Total']
            #result_df.loc[result_df.index[-1], 'Tiempo de intervención'] += df.loc[i, 'Tiempo de intervención'] SOLO COGEMOS EL TUBO A

             #MUCHO CUIDADO CON EL .ILOC PERTENECE A UNA VERSIÓN ANTIGUA DE PANDAS, SI SE ACTUALIZA PANDAS ES POSIBLE QUE DEJE DE FUNCIONAR
    return result_df
################################################################################################################################################

def apply_follow_up_conditions(df, umbral_PDA, umbral_DPR, umbral_Tiempo):
    #Triple condicional para asignar valores a la columna de seguimiento. Los umbrales se definen en el main del codigo
    df['Seguimiento'] = df.apply(
        lambda row: 'SI' if row['Dose_Area_Product_Total'] > umbral_PDA or row['Dose_RP_Total'] > umbral_DPR or row['Tiempo de intervención'] > umbral_Tiempo else 'NO', 
        axis=1
    )
    return df.reset_index(drop=True)

################################################################################################################################################

def export_to_excel(df, excel_file):
    #limpiamos los datos para hacerlos portables más facilmente al excel
    #cambiamos los datos de fecha y hora para que pasen de ser datos tipo "numero" a datos tipo "fecha" y "hora"
    print('Limpiando y añadiendo datos a hoja Excel: '+excel_file)
    
    df['SeriesDate'] = pd.to_datetime(df['SeriesDate'], format='%Y%m%d').dt.date
    df['SeriesTime'] = pd.to_datetime(df['SeriesTime'], format='%H%M%S.%f').dt.time
    df['Dose_RP_Total']= df['Dose_RP_Total'].round(2)
    df['Dose_Area_Product_Total']=df['Dose_Area_Product_Total'].round(2)
    df['Tiempo de intervención']=df['Tiempo de intervención'].round(2)
     
    # Cargar el libro de trabajo existente
    libro = load_workbook(excel_file, keep_vba=True)
    
    # Seleccionar la hoja de trabajo en la que quieres añadir los datos
    hoja_pendientes = libro['Pendientes']
    
    #pegamos los datos a partir de la ultima fila de pacientes rellenada en el excel
   
    ultima_fila_pendientes = 1
    for fila in hoja_pendientes.iter_rows(min_row=1, max_col=1, values_only=True):
            if fila[0] is None:
                break
            ultima_fila_pendientes += 1
        
        # Añadir las filas del DataFrame a partir de la última fila con contenido en la hoja "Pendientes"
    for i, fila in df.iterrows():
            for j, valor in enumerate(fila):
                celda = hoja_pendientes.cell(row=ultima_fila_pendientes + i, column=j + 1, value=valor)
                # Centrar el contenido de la celda
                celda.alignment = Alignment(horizontal='center', vertical='center')
                # Aplicar el formato adecuado si es fecha u hora
                if j == 4:  # Suponiendo que la cuarta columna es la fecha
                    celda.number_format = 'DD/MM/YYYY'  # Formato de fecha
                elif j == 5:  # Suponiendo que la quinta columna es la hora
                    celda.number_format = 'HH:MM'  # Formato de hora
      
        
        # Guardar los cambios en el archivo de Excel
    libro.save(excel_file)
    print("Datos añadidos exitosamente.")
    
###################################################################################################################
###################################################################################################################



###################################################################################################################
###################################################################################################################

def main():
    #fragmento inicial para control de errores
    # Obtener la fecha actual y formatearla para incluirla en el nombre del archivo
    current_date = datetime.now().strftime("%Y-%m-%d")
    log_filename = f'Z:\Dosis pacientes\Alertas_Intervencionismo\_PRUEBAS EN CURSO-AlertInt python\_Logs errores\Biplano\error_log_{current_date}.txt'
    
    # Crear un logger
    logger = logging.getLogger()
    logger.setLevel(logging.ERROR)
    
    # Crear un manejador de archivo que escriba en el archivo log
    file_handler = logging.FileHandler(log_filename)
    file_handler.setLevel(logging.ERROR)
    
    # Definir el formato del log
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    
    # Añadir el manejador al logger
    logger.addHandler(file_handler)

############################################################################
    try:  #ESTE ES EL CODIGO DE VERDAD  
    
        
        #ubicacion de ficheros XML
        input_dir = 'Z:\Dosis pacientes\Alertas_Intervencionismo\Historial pacientes\Biplano'
        print('Buscando pacientes de Biplano en ' + input_dir)
        #hoja de pacietes Alertin
        excel_file = 'Z:\Dosis pacientes\Alertas_Intervencionismo\_PRUEBAS EN CURSO-AlertInt python\ppacientes.xlsm'
        
        #umbrales de notificacion
        umbral_DPR = 5
        umbral_PDA = 500
        umbral_Tiempo = 60
        
        
        all_data = import_xml(input_dir)
    
        if not all_data.empty:
            
            cleaned_data = transform_data(all_data)
            
            filtered_data = remove_zero_rows(cleaned_data)
            
            summed_data = sum_patient_data(filtered_data)
            
            final_data = apply_follow_up_conditions(summed_data,umbral_PDA,umbral_DPR,umbral_Tiempo)
    
            #selección y flitrado de pacienes, solo pega en la hoja de pacientes los que tienes SI en la celda de seguimiento
            df_follow_up = final_data[final_data['Seguimiento'] == 'SI']
            df_follow_up = df_follow_up.reset_index(drop=True)
            print(df_follow_up)
            
            if df_follow_up.empty:
                print('No se encontraron pacientes que necesiten seguimiento')
                sys.exit()
    
            export_to_excel(df_follow_up, excel_file)
            
##############################################################################
        
    except Exception as e:
        # Capturar y registrar el error
        logger.error("Error occurred: %s", e)
        logger.error(traceback.format_exc())
    
    finally:
        # Asegurarse de que se cierra el manejador del archivo
        logger.removeHandler(file_handler)
        file_handler.close()

if __name__ == "__main__":
    main()





