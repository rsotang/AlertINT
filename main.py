import pandas as pd
from openpyxl import load_workbook

# Suponiendo que ya tienes un DataFrame llamado df
df = pd.DataFrame({
    'tiempo': [10.9666667, 1.8833333, 650],
    'Columna2': ['A', 'B', 'C'],
    'Seguimiento':['','' ,'']
})

# Ruta del archivo de Excel
#archivo_excel = 'Pruebas de Victor/pacientesprueba.xlsm'

# Cargar el libro de trabajo existente
#libro = load_workbook(archivo_excel, keep_vba=True)

# Seleccionar la hoja de trabajo en la que quieres añadir los datos
#hoja = libro['Pendientes']

# Encontrar la última fila con contenido en la hoja
#ultima_fila = 1
#for fila in hoja.iter_rows(min_row=1, max_col=1, values_only=True):
    #if all(cell is None for cell in fila):
    #    break
    #ultima_fila += 1

# Añadir las filas del DataFrame a partir de la última fila con contenido
#for i, fila in df.iterrows():
#    for j, valor in enumerate(fila):
#        hoja.cell(row=ultima_fila + i , column=j + 1, value=valor)
#
# Guardar los cambios en el archivo de Excel
#libro.save(archivo_excel)

#print("Datos añadidos exitosamente.")
print(df['tiempo'].dtypes)
df['Seguimiento'] = df['tiempo'].apply(lambda x: 'SI' if x < 60 else 'NO')
print(df)