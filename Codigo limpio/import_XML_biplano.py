
import os
import sys
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import warnings


#esta linea esta para que no se vea el mensaje de error de la libreria pandas al limpiar los datos duplicados del biplano
pd.options.mode.chained_assignment = None  # default='warn'
warnings.simplefilter(action='ignore', category=FutureWarning)


#Esta función intenta leer un archivo XML y devolver un DataFrame con los datos extraídos
#la forma de las funciones esta adecuada a la estructura de los XML del Biplano en concreto
#de acuerdo a la forma de nodos XML padre e hijos.
#
#Para el Biplano se generan archivos XML con todos los pacientes que se han hecho en cada dia
#Esto es por un query automatizado diario.
# 
#El XML tiene una estructura un poco rara por eso mismo.
#Dentro de root hay un único hijo que es <Query> que es en verdad el root que nos interesa
#dentro de Query todos los hijos son de etiqueta <Doseinfo>, que se corresponde con cada paciente
#Dentro de <DoseInfo> (es decir de cada paciente) estan anidados en paralelo los nodos hijos de 
#<ObserverContext>, <CT_Accumulated_Dose_Data> (este nodo está dos veces porque genera uno por panel, por eso de que es un arco biplano y tal), y <CT_Acquisition> (este nodo está tantas veces como veces hayan pisado el pedal de fluoro o cine)
#De momento no parece que haya que modificar el script para adaptarlo al hibrido o a cualquier maquina de Siemenes

###################################################

#METER CODIGOS DE ERROR Y TRY-CATCH 
###################################################

def parse_xml_to_df(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    # Crear una lista para almacenar los datos
    data_list = []

    # Verificar si se encontró <Query_Criteria>
    query_criteria = root.find('Query_Criteria')
    if query_criteria is not None:
        base_data = query_criteria.attrib
        #print("Query_Criteria encontrado:", base_data)

        # Recorrer cada <DoseInfo> y sus subelementos
        for dose_info in query_criteria.findall('DoseInfo'):
            dose_info_data = dose_info.attrib
            #print("DoseInfo encontrado:", dose_info_data)

            # Extraer <Observer_Context> asociado
            observer_context = dose_info.find('Observer_Context')
            if observer_context is not None:
                dose_info_data.update(observer_context.attrib)
                #print("Observer_Context encontrado:", observer_context.attrib)

            # Extraer todos los <CT_Accumulated_Dose_Data>
            for accumulated_dose in dose_info.findall('CT_Accumulated_Dose_Data'):
                accumulated_data = accumulated_dose.attrib
                #print("CT_Accumulated_Dose_Data encontrado:", accumulated_data)
                combined_data = {**base_data, **dose_info_data, **accumulated_data}
                data_list.append(combined_data)

            # Extraer todos los <CT_Acquisition> 
#            for acquisition in dose_info.findall('CT_Acquisition'):
 #               acquisition_data = acquisition.attrib
 #               #print("CT_Acquisition encontrado:", acquisition_data)
 #               combined_data = {**base_data, **dose_info_data, **acquisition_data}
 #               data_list.append(combined_data)
#Esta es la parte que contiene la información mecánica de la maquina y que ene este caso nos da un poco igual
    
    else:
        print("No se encontró <Query_Criteria> en", xml_file)

    # Convertir la lista de diccionarios a un DataFrame
    return pd.DataFrame(data_list) if data_list else pd.DataFrame()

def import_xml(input_dir):
# Leer y combinar todos los archivos XML en un solo DataFrame
    all_data = pd.DataFrame()

    for file in os.listdir(input_dir):
        if file.endswith('.xml'):
            file_path = os.path.join(input_dir, file)
            print("Procesando archivo:", file_path)
            df = parse_xml_to_df(file_path)
            all_data = pd.concat([all_data, df], ignore_index=True)
    
    return all_data

#Esta función esta para eliminar las unidades de los xmls y poder quedarnos solo con los numeros

def clean_and_convert_to_float(column):
        # Extrae solo la parte numérica antes de cualquier carácter no numérico y convierte a float
    return pd.to_numeric(column.str.extract(r'^(\d*\.?\d*)')[0], errors='coerce')



def remove_zero_rows(df):
    return df[(df['Dose_Area_Product_Total'] > 0) & (df['Dose_RP_Total'] > 0)].reset_index(drop=True)

def transform_data(df):
    df['Total_Acquisition_Time'] = clean_and_convert_to_float(df['Total_Acquisition_Time']) if 'Total_Acquisition_Time' in df.columns else 0
    df['Total_Fluoro_Time'] = clean_and_convert_to_float(df['Total_Fluoro_Time']) if 'Total_Fluoro_Time' in df.columns else 0
    df['Dose_Area_Product_Total'] = clean_and_convert_to_float(df['Dose_Area_Product_Total']) * 10000
    df['Dose_RP_Total'] = clean_and_convert_to_float(df['Dose_RP_Total'])
    df['Tiempo de intervención'] = df['Total_Fluoro_Time'] / 60

    df['equipo'] = 'BIPLANO'
    df['servicio'] = 'Neurointervencionismo'
    df['nombre paciente'] = ''
    df['Seguimiento'] = ''

    columns_order = [
        'equipo', 'servicio', 'PatientID', 'nombre paciente', 'SeriesDate', 
        'SeriesTime', 'StudyDescription', 'Dose_Area_Product_Total', 
        'Dose_RP_Total', 'Tiempo de intervención', 'Seguimiento'
    ]
    return df.reindex(columns=columns_order)

def aggregate_patient_data(df):
    # Agrupar por 'PatientID' y sumar las columnas numéricas
    aggregated_df = df.groupby('PatientID', as_index=False).agg({
        'SeriesDate': 'first',  # Puede usar 'first' o 'min' si la fecha es la misma para todas las entradas
        'SeriesTime': 'first',  # Similar a la fecha, depende de su uso
        'StudyDescription': 'first',
        'Tiempo de intervención': 'sum',
        'Dose_Area_Product_Total': 'sum',
        'Dose_RP_Total': 'sum'
    })
    return aggregated_df

def sum_patient_data(df):
    result_df = pd.DataFrame(columns=df.columns)
    for i in range(len(df)):
        if i == 0 or df.loc[i, 'PatientID'] != df.loc[i-1, 'PatientID']:
            result_df = pd.concat([result_df, df.loc[[i]]], ignore_index=True)
        else:
            result_df.loc[result_df.index[-1], 'Dose_Area_Product_Total'] += df.loc[i, 'Dose_Area_Product_Total']
            result_df.loc[result_df.index[-1], 'Dose_RP_Total'] += df.loc[i, 'Dose_RP_Total']
            result_df.loc[result_df.index[-1], 'Tiempo de intervención'] += df.loc[i, 'Tiempo de intervención']
    return result_df

def apply_follow_up_conditions(df, umbral_PDA, umbral_DPR, umbral_Tiempo):
    df['Seguimiento'] = df.apply(
        lambda row: 'SI' if row['Dose_Area_Product_Total'] > umbral_PDA or row['Dose_RP_Total'] > umbral_DPR or row['Tiempo de intervención'] > umbral_Tiempo else 'NO', 
        axis=1
    )
    return df.reset_index(drop=True)


def export_to_excel(df, excel_file):
    df['SeriesDate'] = pd.to_datetime(df['SeriesDate'], format='%Y%m%d').dt.date
    df['SeriesTime'] = pd.to_datetime(df['SeriesTime'], format='%H%M%S.%f').dt.time
    
     
    # Cargar el libro de trabajo existente
    libro = load_workbook(excel_file, keep_vba=True)
    
    # Seleccionar la hoja de trabajo en la que quieres añadir los datos
    hoja_pendientes = libro['Pendientes']

    ultima_fila_pendientes = len(list(hoja_pendientes.iter_rows(min_row=1, max_col=1, values_only=True)))

    for i, fila in df.iterrows():
        for j, valor in enumerate(fila):
            celda = hoja_pendientes.cell(row=ultima_fila_pendientes + i, column=j + 1, value=valor)
            celda.alignment = Alignment(horizontal='center', vertical='center')
            if j == 4:
                celda.number_format = 'DD/MM/YYYY'
            elif j == 5:
                celda.number_format = 'HH:MM'

    ultima_fila_pendientes = 1
    for fila in hoja_pendientes.iter_rows(min_row=1, max_col=1, values_only=True):
            if fila[0] is None:
                break
            ultima_fila_pendientes += 1
        
        # Añadir las filas del DataFrame a partir de la última fila con contenido en la hoja "Pendientes"
    for i, fila in df.iterrows():
            for j, valor in enumerate(fila):
                celda = hoja_pendientes.cell(row=ultima_fila_pendientes + i, column=j + 1, value=valor)
                # Centrar el contenido de la celda
                celda.alignment = Alignment(horizontal='center', vertical='center')
                # Aplicar el formato adecuado si es fecha u hora
                if j == 4:  # Suponiendo que la cuarta columna es la fecha
                    celda.number_format = 'DD/MM/YYYY'  # Formato de fecha
                elif j == 5:  # Suponiendo que la quinta columna es la hora
                    celda.number_format = 'HH:MM'  # Formato de hora
      
        
        # Guardar los cambios en el archivo de Excel
    libro.save(excel_file)
    print("Datos añadidos exitosamente.")
    
#############################################

def main():
    input_dir = 'C:/Users/53112727T/Desktop/Alertin 1.1/Pruebas de Victor/Paciente prueba Alertint/XML Biplano'
    excel_file = 'ppacientes.xlsm'
    umbral_DPR = 5
    umbral_PDA = 500
    umbral_Tiempo = 10
    all_data = import_xml(input_dir)

    if not all_data.empty:
        
        cleaned_data = transform_data(all_data)
        
        filtered_data = remove_zero_rows(cleaned_data)
        
        summed_data = sum_patient_data(filtered_data)
        
        final_data = apply_follow_up_conditions(summed_data,umbral_PDA,umbral_DPR,umbral_Tiempo)

        df_follow_up = final_data[final_data['Seguimiento'] == 'SI']
        df_follow_up = df_follow_up.reset_index(drop=True)
        print(df_follow_up)
        
        if df_follow_up.empty:
            print('No se encontraron pacientes que necesiten seguimiento')
            sys.exit()

        export_to_excel(df_follow_up, excel_file)


if __name__ == "__main__":
    main()





