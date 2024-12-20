#Este script está diseñado para extraer información de dosis util para nuestra hoja de pacientes.xlsx que ya existe en AlertINT de archivos SR
#De momento solo he comprobado que funciona para las Phillips, hay que comprobar con otros fabricantes.

#LA UTILIDAD REAL, ES QUE SI SE ROMPE EL CAREANALYTICS O LO QUE SEA, CON ESTA HERRAMIENTA SE PUEDE SEGUIR FUNCIONANDO EN ALERTIN, SOLO SE NECESITAN LOS SR QUE GENERAN LAS MAQUINAS AUTOMATICAMENTE

#Puede haber problemas de lectura de los archivos porque las plantillas de los SR no están del todo estandarizados.

#De todas formas el codigo está diseñado buscando el codigo numerico NEMA de las etiquetas que se necesitan, que eso si que está estandarizado.asi que de primeras podría ser robusto

#tengo que añadir todavia codigos de errores asi que en esos codigos puedo meter alguna excepción para que busque con otro codigo NEMA redundante que tenga la misma información.

#El script no es nada especial, simplemente es un bucle de busqueda de datos teniendo en cuenta las peculiaridades del formalismo SR. 

#Los SR son esencialmente XML complicados, no solo están en estructura anidada, si no que cada nido tiene una estructura de cabecera y etiquetas con información, asi que es como un nido de arrays de arrays. (se entiende mejor en la pagina de la NEMA)

#En nuestro caso estamos tabajando con un estandar SR en el que los datos están organizados siguiento la estructura definida por la NEMA como TID 10001 Projection X-Ray Radiation Dose TID 10002 y TID10003

#En el caso de la angiografia el SR es esncialmente una cabecera con la información del paciente y la maquina y un resumen del contenido.

#Despues empiezan los nidos, que son los "eventos de radiación" que es cada vez que se ha pisado el pedal

#Cada "evento" se desnida en un nido con la información de la dosis, un nido con la información mecanica y otro nido con la información de imagen, en la que cada frame a su vez es otro subnido (es un lio importante)

#por suerte la información relevante para nosotros es relativamente accesible y esta al principio del archivo, por lo que el script corre bastante rapido, si no hay que tener en cuenta que cada SR pueden ser entre 10^5 y 10^7 lineas de texto que recorrer buscando un dato, es decir que el codigo es sencillo y corre rapido en esta situación, pero para tareas mas complejas habría que optimizarlo


import pydicom
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Función para extraer la información de dosis de un archivo DICOM SR
def extraer_datos_dosis(archivo_dicom):
    ds = pydicom.dcmread(archivo_dicom)
    
    print('Importando archivos')
    
    dosis_total = 0.0
    tiempo_total_intervencion = 0.0
    pda_total = 0.0
    tiempo_fluo = 0.0  #Son datos separados que no se juntán en ninguna etiqueta, 
    tiempo_adq = 0.0   #asi que los sumamos nosostros a mano
    
    # Navegar a través de los eventos de irradiación
    for item in ds.ContentSequence:
        if hasattr(item, 'ContentSequence'):
            for subitem in item.ContentSequence:
                # Extraer la dosis a punto de referencia
                if subitem.ConceptNameCodeSequence[0].CodeValue == '113725':  # Código NEMA para dosis a punto de referencia
                    dosis_referencia = float(subitem.MeasuredValueSequence[0].NumericValue)
                    dosis_total += dosis_referencia

                # Extraer el tiempo de fluoroscopia
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113730':  # Código NEMA para tiempo de fluoroscopia
                    tiempo_fluo = float(subitem.MeasuredValueSequence[0].NumericValue)

                # Extraer el tiempo de adquisición
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113855':  # Código NEMA para tiempo de adquisición
                    tiempo_adq = float(subitem.MeasuredValueSequence[0].NumericValue)

                # Extraer el producto dosis-área (PDA)
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113722':  # Código NEMA para producto dosis-área
                    pda = float(subitem.MeasuredValueSequence[0].NumericValue)
                    pda_total += pda*10000

    # Calcular el tiempo total de intervención fuera del bucle
    tiempo_total_intervencion = tiempo_fluo/60 #tiempo_adq + tiempo_fluo 

    return dosis_total, tiempo_total_intervencion, pda_total

#######################################################################

# Ruta de la carpeta que contiene los archivos SRDICOM
carpeta_srdicom = 'Pruebas de Victor/Paciente prueba Alertint/SR Phillips'

#####################################################################
# Lista para almacenar la información de cada archivo
datos_pacientes = []
print('procesando datos')
# Iterar sobre los archivos en la carpeta
for archivo in os.listdir(carpeta_srdicom):
    if archivo.endswith('.dcm'):
        archivo_dicom = os.path.join(carpeta_srdicom, archivo)
        ds = pydicom.dcmread(archivo_dicom)

        # Extraer información adicional del paciente y estudio
        id_paciente = ds.PatientID
        nombre_paciente = str(ds.PatientName)
        fecha_intervencion = ds.StudyDate
        hora_intervencion = ds.StudyTime
        descripcion_estudio = ds.StudyDescription if 'StudyDescription' in ds else 'N/A'

        # Extraer información de dosis, tiempo y PDA
        dosis_pref, tiempo_total_intervencion, pda_total = extraer_datos_dosis(archivo_dicom)

        # Añadir los datos al DataFrame
        datos_pacientes.append({
            'equipo': 'Hemodinámicas Philips',
            'servicio': 'Cardiología',
            'PatientID': id_paciente,
            'nombre paciente': nombre_paciente,
            'SeriesDate': fecha_intervencion,
            'SeriesTime': hora_intervencion,
            'StudyDescription': descripcion_estudio,
            'Dose_Area_Product_Total': pda_total,
            'Dose_RP_Total': dosis_pref,
            'Tiempo de intervención': tiempo_total_intervencion, 
            'Seguimiento': ''
        })

# Crear un DataFrame con los datos recolectados
df = pd.DataFrame(datos_pacientes)




#####################################################################

print('limpiando datos')



df['SeriesDate'] = pd.to_datetime(df['SeriesDate'], format='%Y%m%d').dt.date

df['SeriesTime'] = pd.to_datetime(df['SeriesTime'], format='%H%M%S.%f').dt.time
print(df)

  #Condiciones de seguimiento
######################################
#################################

umbral_PDA = 500 #Gycm2
umbral_DPR = 5 #Gy
umbral_Tiempo = 10 #min

################################
#######################################

df['Seguimiento'] = df.apply(lambda row: 'SI' if row['Dose_Area_Product_Total'] > umbral_PDA or row['Dose_RP_Total'] > umbral_DPR or row['Tiempo de intervención'] > umbral_Tiempo else 'NO', axis=1) 
result_df = df

 #Generamos array secundario para seleccionar los pacientes con seguimieno
result_df = df[df['Seguimiento'] == 'SI']
result_df = result_df.reset_index(drop=True)
print(result_df)
df = result_df
# Ruta del archivo de Excel

if df.empty:
            print('No se encontraron pacientes que necesiten seguimiento')
            sys.exit()
print('moviendo al excel')

#############################################################
archivo_excel = 'ppacientes.xlsm'
#############################################################

# Cargar el libro de trabajo existente
libro = load_workbook(archivo_excel, keep_vba=True)

# Seleccionar la hoja de trabajo en la que quieres añadir los datos
hoja_pendientes = libro['Pendientes']



# Encontrar la última fila con contenido en la hoja "Pendientes"
ultima_fila_pendientes = 1
for fila in hoja_pendientes.iter_rows(min_row=1, max_col=1, values_only=True):
    if fila[0] is None:
        break
    ultima_fila_pendientes += 1

# Añadir las filas del DataFrame a partir de la última fila con contenido en la hoja "Pendientes"
for i, fila in df.iterrows():
    for j, valor in enumerate(fila):
        celda = hoja_pendientes.cell(row=ultima_fila_pendientes + i, column=j + 1, value=valor)
        # Centrar el contenido de la celda
        celda.alignment = Alignment(horizontal='center', vertical='center')
        # Aplicar el formato adecuado si es fecha u hora
        if j == 4:  # Suponiendo que la cuarta columna es la fecha
            celda.number_format = 'DD/MM/YYYY'  # Formato de fecha
        elif j == 5:  # Suponiendo que la quinta columna es la hora
            celda.number_format = 'HH:MM'  # Formato de hora

  



# Guardar los cambios en el archivo de Excel
libro.save(archivo_excel)

print("Datos añadidos exitosamente.")
########################################################


#Hay que investigar las UID Dicom para poder distinguir las maquinas entre si y añadir las columnas que nos faltan.