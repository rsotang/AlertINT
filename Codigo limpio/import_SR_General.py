import pydicom
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment


#ESTA ES UNA VERSION REPETIDA DEL SCRIPT SR_PHILLIPS.PY QUE ES EL ORIGINAL DONDE ESTÁN TODOS LOS COMENTARIOS

#De momento parece que para General y Phillips los campos de los que sacamos los datos coinciden, pero puede ser susceptible de error, asi que en caso de que no se importen bien los datos de una maquina solo un posible erro rpuede ser que hayan actualizado los campos del SR de esa maquina y que no lo coja bien el script.

def extraer_datos_dosis(archivo_dicom):
    ds = pydicom.dcmread(archivo_dicom)

    dosis_total = 0.0
    tiempo_total_intervencion = 0.0
    pda_total = 0.0
    tiempo_fluo = 0.0  #Son datos separados que no se juntán en ninguna etiqueta, 
    tiempo_adq = 0.0   #asi que los sumamos nosostros a mano

    # Navegar a través de los eventos de irradiación
    for item in ds.ContentSequence:
        if hasattr(item, 'ContentSequence'):
            for subitem in item.ContentSequence:
                # Extraer la dosis a punto de referencia
                if subitem.ConceptNameCodeSequence[0].CodeValue == '113725':  # Código NEMA para dosis a punto de referencia
                    dosis_referencia = float(subitem.MeasuredValueSequence[0].NumericValue)
                    dosis_total += dosis_referencia

                # Extraer el tiempo de fluoroscopia
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113730':  # Código NEMA para tiempo de fluoroscopia
                    tiempo_fluo = float(subitem.MeasuredValueSequence[0].NumericValue)

                # Extraer el tiempo de adquisición
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113855':  # Código NEMA para tiempo de adquisición
                    tiempo_adq = float(subitem.MeasuredValueSequence[0].NumericValue)

                # Extraer el producto dosis-área (PDA)
                elif subitem.ConceptNameCodeSequence[0].CodeValue == '113722':  # Código NEMA para producto dosis-área
                    pda = float(subitem.MeasuredValueSequence[0].NumericValue)
                    pda_total += pda*10000

    # Calcular el tiempo total de intervención fuera del bucle
    tiempo_total_intervencion = tiempo_fluo/60 #tiempo_adq + tiempo_fluo 

    return dosis_total, tiempo_total_intervencion, pda_total

# Ruta de la carpeta que contiene los archivos SRDICOM
carpeta_srdicom = 'Pruebas de Victor/Paciente prueba Alertint/SR General'

#####################################################################
# Lista para almacenar la información de cada archivo
datos_pacientes = []
print('procesando datos')
# Iterar sobre los archivos en la carpeta
for archivo in os.listdir(carpeta_srdicom):
    if archivo.endswith('.dcm'):
        archivo_dicom = os.path.join(carpeta_srdicom, archivo)
        ds = pydicom.dcmread(archivo_dicom)

        # Extraer información adicional del paciente y estudio
        id_paciente = ds.PatientID
        nombre_paciente = str(ds.PatientName)
        fecha_intervencion = ds.StudyDate
        hora_intervencion = ds.StudyTime
        descripcion_estudio = ds.StudyDescription if 'StudyDescription' in ds else 'N/A'

        # Extraer información de dosis, tiempo y PDA
        dosis_pref, tiempo_total_intervencion, pda_total = extraer_datos_dosis(archivo_dicom)

        # Añadir los datos al DataFrame
        datos_pacientes.append({
            'equipo': 'GE Vascular',
            'servicio': 'Cardiología',
            'PatientID': id_paciente,
            'nombre paciente': nombre_paciente,
            'SeriesDate': fecha_intervencion,
            'SeriesTime': hora_intervencion,
            'StudyDescription': descripcion_estudio,
            'Dose_Area_Product_Total': pda_total,
            'Dose_RP_Total': dosis_pref,
            'Tiempo de intervención': tiempo_total_intervencion, 
            'Seguimiento': ''
        })

# Crear un DataFrame con los datos recolectados
df = pd.DataFrame(datos_pacientes)




#####################################################################

print('limpiando datos')



df['SeriesDate'] = pd.to_datetime(df['SeriesDate'], format='%Y%m%d').dt.date

df['SeriesTime'] = pd.to_datetime(df['SeriesTime'], format='%H%M%S.%f').dt.time
print(df)

  #Condiciones de seguimiento
######################################
#################################

umbral_PDA = 500 #Gycm2
umbral_DPR = 5 #Gy
umbral_Tiempo = 10 #min

################################
#######################################

df['Seguimiento'] = df.apply(lambda row: 'SI' if row['Dose_Area_Product_Total'] > umbral_PDA or row['Dose_RP_Total'] > umbral_DPR or row['Tiempo de intervención'] > umbral_Tiempo else 'NO', axis=1) 
result_df = df

 #Generamos array secundario para seleccionar los pacientes con seguimieno
result_df = df[df['Seguimiento'] == 'SI']
result_df = result_df.reset_index(drop=True)
print(result_df)
# Ruta del archivo de Excel
df = result_df
if df.empty:
            print('No se encontraron pacientes que necesiten seguimiento')
            sys.exit()
print('moviendo al excel')

#############################################################
archivo_excel = 'ppacientes.xlsm'
#############################################################

# Cargar el libro de trabajo existente
libro = load_workbook(archivo_excel, keep_vba=True)

# Seleccionar la hoja de trabajo en la que quieres añadir los datos
hoja_pendientes = libro['Pendientes']



# Encontrar la última fila con contenido en la hoja "Pendientes"
ultima_fila_pendientes = 1
for fila in hoja_pendientes.iter_rows(min_row=1, max_col=1, values_only=True):
    if fila[0] is None:
        break
    ultima_fila_pendientes += 1

# Añadir las filas del DataFrame a partir de la última fila con contenido en la hoja "Pendientes"
for i, fila in df.iterrows():
    for j, valor in enumerate(fila):
        celda = hoja_pendientes.cell(row=ultima_fila_pendientes + i, column=j + 1, value=valor)
        # Centrar el contenido de la celda
        celda.alignment = Alignment(horizontal='center', vertical='center')
        # Aplicar el formato adecuado si es fecha u hora
        if j == 4:  # Suponiendo que la cuarta columna es la fecha
            celda.number_format = 'DD/MM/YYYY'  # Formato de fecha
        elif j == 5:  # Suponiendo que la quinta columna es la hora
            celda.number_format = 'HH:MM'  # Formato de hora

  



# Guardar los cambios en el archivo de Excel
libro.save(archivo_excel)

print("Datos añadidos exitosamente.")
#Hay que añadir el filtrado de datos y la exportación a excel